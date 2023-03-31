from ast import Break
import sys
import os
from tracemalloc import stop
import openpyxl
from openpyxl.styles import PatternFill
from PIL import Image
import pyautogui
import win32gui, win32con, time

VerticalSize = 500
HorizontalSize = 500 #max is 676

folder_path = os.path.abspath(os.path.join(sys.argv[0],'../')) 

#Get path to images
photos_path = folder_path + '/photos2'

for root, dirs, files in os.walk(r''+ photos_path):
    photos_files = files
    break

#if the program is paused while processing, to continue where you were, put in progress how many files you processed
progress = 0
for d in range(progress):
    photos_files.pop(0)

for f in photos_files:
    
    file = (photos_path + '/' + f)

    #Resize images
    image = Image.open(file)
    resized_image = image.resize((HorizontalSize, VerticalSize))

    #Get every pixel color
    photo_colors = []
    for i in range(HorizontalSize):
        for c in range(VerticalSize):
            photo_colors.append('%02x%02x%02x' % resized_image.getpixel((i,c))) #get color and change to hex
    
    #Color excel
    wb = openpyxl.load_workbook(folder_path + '/film.xlsx')
    ws = wb.active

    Letter = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T', 'U', 'V','W','X','Y','Z']
    count =  0
    l = 1 #Used to change AZ to BA / BZ to CA...

    for i in range(HorizontalSize):
        for c in range(VerticalSize):

            color = photo_colors[count]
            count += 1

            #Convert Letter into numbers
            if i >= len(Letter):
                column = Letter[l - 1] + Letter[i - len(Letter) * l]
            else: 
                column = Letter[i]
            if c == VerticalSize - 1 and column == Letter[l - 1] + 'Z':
                 l += 1

            ws[ column + str(c + 1) ].fill = PatternFill(start_color = color, fill_type = 'solid') #set color to cell

    wb.save(folder_path + '/output/' + 'frame.xlsx') #change this to don't overwrite your last excel

    #os.startfile(folder_path + '/output/'  + 'frame.xlsx') #open
    #time.sleep(2)                                          #you can time this, but it will probably eventually fail
    #hwnd = win32gui.GetForegroundWindow()
    #win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
    #time.sleep(3)
    #pyautogui.screenshot(region=(0,0, 1920, 1020)).save(r'' + folder_path  + '/screenshots/' + f + '.png') #screenshot to excel
    #os.system("TASKKILL /F /IM excel.exe")                 #close every excel file
    
    progress += 1
    
    #stats
    print('Progress: ' + str(progress) + ' / ' + str(len(photos_files)))
    print('Percentage: ' + str(round(progress / len(photos_files) * 100, 2)) + '%')
    print(f)